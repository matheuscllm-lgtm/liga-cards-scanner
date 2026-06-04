"""Geracao do relatorio XLSX com formatacao."""
from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from src.matching.card_matcher import Comparison

HEADER_FILL = "FF003366"
APPROVED_FILL = "FFDCEFDC"
REJECTED_FILL = "FFFCE4E4"

HUMAN_LABELS = {
    "card_name": "Card",
    "set_name": "Set",
    "price_liga_brl": "Liga R$",
    "price_tcg_usd": "TCG USD",
    "price_tcg_brl": "TCG R$",
    "margin_percent": "Margem %",
    "exchange_rate": "Cambio",
    "liga_url": "Link Liga",
    "tcg_url": "Link TCG",
    "status": "Status",
    "match_score": "Score",
}

CURRENCY_BRL_COLS = {"price_liga_brl", "price_tcg_brl"}
CURRENCY_USD_COLS = {"price_tcg_usd"}
PERCENT_COLS = {"margin_percent"}
URL_COLS = {"liga_url", "tcg_url"}


def write_xlsx(items: list["Comparison"], path: Path) -> None:
    """Gera planilha XLSX com cabecalho colorido, formatacao numerica e
    linhas tingidas por status. Importa openpyxl tardiamente para nao
    pesar nos modos que so usam CSV/JSON."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Deals"

    if not items:
        ws.cell(row=1, column=1, value="Nenhum match")
        wb.save(path)
        return

    fields = list(asdict(items[0]).keys())
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color="FFFFFFFF")
    approved_fill = PatternFill("solid", fgColor=APPROVED_FILL)
    rejected_fill = PatternFill("solid", fgColor=REJECTED_FILL)

    for col_idx, key in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=HUMAN_LABELS.get(key, key))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, item in enumerate(items, start=2):
        row = asdict(item)
        fill = approved_fill if row.get("status") == "approved" else rejected_fill
        for col_idx, key in enumerate(fields, start=1):
            value = row[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            if key in CURRENCY_BRL_COLS:
                cell.number_format = 'R$ #,##0.00'
            elif key in CURRENCY_USD_COLS:
                cell.number_format = '"US$ "#,##0.00'
            elif key in PERCENT_COLS:
                cell.number_format = '0.00"%"'
            elif key in URL_COLS and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="FF0563C1", underline="single")

    # Autosize aproximado: pega a maior largura por coluna.
    for col_idx, key in enumerate(fields, start=1):
        max_len = len(str(HUMAN_LABELS.get(key, key)))
        for item in items:
            v = asdict(item).get(key)
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
